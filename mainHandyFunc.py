# for Logging
import logging
import warnings

import json
import pandas as pd
import openpyxl

# for path processing
import os

import openpyxl.styles as styl


def getWhereIsMain():
    if os.path.exists(os.path.join(os.getcwd(), 'pythumb_base_path.txt')):
        with open(os.path.join(os.getcwd(), 'pythumb_base_path.txt'), 'r', encoding='utf-16') as file:
            base_path = file.read().replace('\n', '')
    else:
        base_path = os.getcwd()
    return base_path


def get_file_info():

    base_path = getWhereIsMain()

    # get main file name from setting json
    with open(os.path.join(base_path, '0_setting', 'setting.json'), encoding='utf-8') as json_file:
        data = json.load(json_file)

    logging.info('Main File -- ' + os.path.join(base_path, data['main_file']))

    wb = openpyxl.load_workbook(os.path.join(
        base_path, data['main_file']))
    sheet_go = wb['GO']
    data_dict = {
        'file_path': [],
        'file_name': [],
        'output_name': [],
        'time': [],
    }
    for idx in range(7, sheet_go.max_row):
        if sheet_go['E'+str(idx)].value is not None:
            data_dict['file_path'].append(sheet_go['E'+str(idx)].value)
            data_dict['file_name'].append(sheet_go['F'+str(idx)].value)
            if not sheet_go['G'+str(idx)].value in data_dict['output_name']:
                data_dict['output_name'].append(sheet_go['G'+str(idx)].value)
            else:
                out_name = str(idx)+'_'+str(sheet_go['G'+str(idx)].value)
                data_dict['output_name'].append(out_name)
            data_dict['time'].append(sheet_go['H'+str(idx)].value)

    return pd.DataFrame.from_dict(data_dict)


def changeCl_df(mapping_df, in_df):
    sub_df = pd.DataFrame()
    for cl in mapping_df.columns.tolist():
        if cl == 'FLAG' and 'FLAG' in in_df.columns.tolist():
            sub_df['FLAG'] = in_df['FLAG']
        else:
            if mapping_df[cl][0] != '':
                if mapping_df[cl][0] in in_df.columns.tolist():

                    sub_df[cl] = in_df[mapping_df[cl][0]]

            else:
                sub_df[cl] = ''

    sub_df.fillna('')

    return sub_df


def readMapping(mediaType='SS'):
    mainFolder_path = getWhereIsMain()

    # get main file name from setting json
    with open(os.path.join(mainFolder_path, '0_driver', 'setting.json'), encoding='utf-8') as json_file:
        setting_json = json.load(json_file)

    filePath = os.path.join(
        mainFolder_path, '0_driver', setting_json['columnMapping'])

    mapping_df = pd.read_excel(
        filePath, sheet_name='adsInfoCLMapping', header=0, engine='openpyxl')
    mapping_df = mapping_df.fillna('')
    mapping_df = mapping_df[mapping_df['FLAG']
                            == mediaType].reset_index(drop=True)

    mapping_df.fillna('')

    return mapping_df


def fommatExcelFile(filePath):
    # def style
    allCell_style = styl.NamedStyle(name="allCell_style")
    allCell_style.font = styl.Font(name="メイリオ", size=8)
    allCell_style.alignment = styl.Alignment(
        horizontal="left", vertical="center")

    # open file
    wb1 = openpyxl.load_workbook(filePath)
    for ws in wb1.worksheets:
        for row in ws:
            for cell in row:
                cell.style = allCell_style  # apply style to each cell

    # save wb
    wb1.save(filePath)
    wb1.close()

    """
    # %%
    theURL = 'https://ads.twitter.com/data/composer/18ce53vzsrq/api/tweet?as_user_id=2517371725&text=python-test-1216-1539&card_uri=card%3A%2F%2F1468903518165954560&nullcast=true'
    needed_cookie_array = ['_ga', '_gid', '_twitter_sess', 'att', 'auth_token', 'ct0', 'gt',
                            'guest_id', 'guest_id_ads', 'guest_id_marketing', 'kdt', 'lang', 'personalization_id', 'twid']
    the_cookies = {}
    for ck in needed_cookie_array:
        the_cookies[ck] = firefox.get_cookie(ck)['value']

    response = requests.post(theURL, headers=the_headers, cookies=the_cookies)
    print(response.status_code)

    # %%
    import itertools
    ck_comp_list = []
    needed_cookie_array = ['_ga', '_gid', '_twitter_sess', 'att', 'auth_token', 'ct0', 'gt',
                            'guest_id', 'guest_id_ads', 'guest_id_marketing', 'personalization_id']

    for i in range(3, len(needed_cookie_array)):
        comp_ojb = itertools.combinations(needed_cookie_array, i)
        ck_comp_list += list(comp_ojb)

    final_comp = []
    for ck_comp in ck_comp_list:
        the_cookies = {}
        for ck in ck_comp:
            the_cookies[ck] = firefox.get_cookie(ck)['value']

        response = requests.post(
            theURL, headers=the_headers, cookies=the_cookies)
        if response.status_code != 403:
            final_comp.append(ck_comp)
            break
        else:
            print('{} - {}'.format(response.status_code, ck_comp))
        sleep(3)

    the_data = {
        'as_user_id': '2517371725',
        'text': 'python-test_2021/12/16-maruchi',
        'scheduled_at': '2021-12-17T10:57:17Z',
        'media_ids': '1469652676216782851',
        'nullcast': 'nullcast'
    }

    script_text = '''var xmlHttp = new XMLHttpRequest();
    theUrl = 'https://ads.twitter.com/data/composer/18ce53vzsrq/api/tweet?as_user_id=2517371725&text=python-test-1217-1654&card_uri=card%3A%2F%2F1468903518165954560&nullcast=true';
    thepayLoad ='{}'
    xmlHttp.open("POST", theUrl, false);
    xmlHttp.setRequestHeader("Content-Type", "application/json;charset=UTF-8");
    xmlHttp.send(thepayLoad);
    return xmlHttp.responseText'''

    result = firefox.execute_script(script_text)
    result_json = json.loads(result)
    """


def get_cr_folder(cr_path):
    # get all file in CR_folder
    # return as an array
    logging.info('FUNC(login): get_cr_folder')
    out_array = []
    for path, subdirs, files in os.walk(cr_path):
        for name in files:
            # logging.info(os.path.join(path,name))
            out_array.append([name, os.path.join(path, name)])
    return out_array
